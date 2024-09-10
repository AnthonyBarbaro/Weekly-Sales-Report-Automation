import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

# Function to process each file (MV and LM) and return the processed DataFrame
def process_file(file_path):
    if not os.path.exists(file_path):
        print(f"Error: The file at path {file_path} does not exist.")
        return None

    # Load the Excel file into a DataFrame, specifying the correct header row
    df = pd.read_excel(file_path, header=4)

    # Strip any leading/trailing spaces from column names and convert to lower case
    df.columns = df.columns.str.strip().str.lower()

    # Rename columns to handle unnamed columns properly
    df.columns = [
        "order id", "order time", "budtender name", "customer name", "customer type", 
        "vendor name", "product name", "category", "package id", "batch id", 
        "external package id", "total inventory sold", "unit weight sold", "total weight sold", 
        "gross sales", "inventory cost", "discounted amount", "loyalty as discount", 
        "net sales", "return date", "upc gtin (canada)", "provincial sku (canada)", 
        "producer", "order profit"
    ]

    # Convert 'order time' to datetime and create 'day of week' column
    df['order time'] = pd.to_datetime(df['order time'])
    df['day of week'] = df['order time'].apply(lambda date: date.strftime('%A'))

    return df

# Load both MV and LM files
mv_data = process_file('salesMV.xlsx')
lm_data = process_file('salesLM.xlsx')

# Define the discount and kickback logic
def apply_discounts_and_kickbacks(data, discount, kickback):
    data['discount amount'] = data['gross sales'] * discount
    data['kickback amount'] = data['inventory cost'] * kickback
    return data

# Define the criteria for each brand
brand_criteria = {
    #'BTC Ventures': {
    #    'days': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'],
    #    'discount': 0.50,
    #    'kickback': 0.25,
    #    'categories': ['Concentrate']
    #},
    'Med For America Inc.': {
        'days': ['Monday'],
        'discount': 0.50,
        'kickback': 0.25
    },
    'Kiva': {
        'vendors': ['KIVA / LCISM CORP', 'Vino & Cigarro, LLC'],
        'days': ['Monday'],
        'discount': 0.50,
        'kickback': 0.25,
        'brands': ['Terra', 'Petra', 'Kiva', 'Lost Farms', 'Camino']
    },
    'Big Petes': {
        'vendors': ['KIVA / LCISM CORP', 'Vino & Cigarro, LLC'],
        'days': ['Tuesday'],
        'discount': 0.50,
        'kickback': 0.25,
        'brands': ['Big Pete']
    },
    'Garden Of Weeden Inc.': {
        'days': ['Sunday', 'Friday', 'Saturday'],
        'discount': 0.50,
        'kickback': 0.25,
        'brands': ['Huxley','Wav']
    }
}

# Directory to save individual brand reports
output_dir = 'brand_reports'
os.makedirs(output_dir, exist_ok=True)


# Process each brand for both MV and LM data
consolidated_summary = []

for brand, criteria in brand_criteria.items():
    # Filter MV data
    if "vendors" in criteria:
        mv_brand_data = mv_data[(mv_data['vendor name'].isin(criteria['vendors'])) & (mv_data['day of week'].isin(criteria['days']))].copy()
        lm_brand_data = lm_data[(lm_data['vendor name'].isin(criteria['vendors'])) & (lm_data['day of week'].isin(criteria['days']))].copy()
    else:
        mv_brand_data = mv_data[(mv_data['vendor name'] == brand) & (mv_data['day of week'].isin(criteria['days']))].copy()
        lm_brand_data = lm_data[(lm_data['vendor name'] == brand) & (lm_data['day of week'].isin(criteria['days']))].copy()

    # Additional filtering for categories or product names
    if 'categories' in criteria:
        mv_brand_data = mv_brand_data[mv_brand_data['category'].isin(criteria['categories'])]
        lm_brand_data = lm_brand_data[lm_brand_data['category'].isin(criteria['categories'])]
    if 'brands' in criteria:
        mv_brand_data = mv_brand_data[mv_brand_data['product name'].apply(lambda x: any(b in x for b in criteria['brands']))]
        lm_brand_data = lm_brand_data[lm_brand_data['product name'].apply(lambda x: any(b in x for b in criteria['brands']))]

    if not mv_brand_data.empty or not lm_brand_data.empty:
        # Apply discounts and kickbacks
        mv_brand_data = apply_discounts_and_kickbacks(mv_brand_data, criteria['discount'], criteria['kickback'])
        lm_brand_data = apply_discounts_and_kickbacks(lm_brand_data, criteria['discount'], criteria['kickback'])

        # Calculate the date range from the 'Order Time' column
        start_date = mv_brand_data['order time'].min().strftime('%Y-%m-%d')
        end_date = mv_brand_data['order time'].max().strftime('%Y-%m-%d')
        date_range = f"{start_date}_to_{end_date}"

        # Create summary for each location
        mv_summary = mv_brand_data.agg({
            'gross sales': 'sum',
            'inventory cost': 'sum',
            'discount amount': 'sum',
            'kickback amount': 'sum'
        }).to_frame().T
        mv_summary['location'] = 'MV'

        lm_summary = lm_brand_data.agg({
            'gross sales': 'sum',
            'inventory cost': 'sum',
            'discount amount': 'sum',
            'kickback amount': 'sum'
        }).to_frame().T
        lm_summary['location'] = 'LM'

        # Consolidated summary for this brand
        brand_summary = pd.concat([mv_summary, lm_summary])
        brand_summary['brand'] = brand
        brand_summary['days active'] = ', '.join(criteria['days'])
        consolidated_summary.append(brand_summary)

        # Save individual brand report with date range in filename
        output_filename = f"{output_dir}/{brand.replace('/', ' ')}_report_{date_range}.xlsx"
        with pd.ExcelWriter(output_filename) as writer:
            mv_brand_data.to_excel(writer, sheet_name='MV_Sales', index=False)
            lm_brand_data.to_excel(writer, sheet_name='LM_Sales', index=False)
            brand_summary.to_excel(writer, sheet_name='Summary', index=False)

        # Apply formatting using openpyxl
        workbook = load_workbook(output_filename)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
                        # Apply row height adjustment
            for column in sheet.columns:
                max_length = max([len(str(cell.value)) for cell in column])
                adjusted_width = max_length
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width + 1
            for row in sheet.iter_rows():
                sheet.row_dimensions[row[0].row].height = 17  # Adjust the value to your desired height
            for cell in sheet["1:1"]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
        workbook.save(output_filename)

# Create a consolidated report for all brands with date range in filename
if consolidated_summary:
    consolidated_df = pd.concat(consolidated_summary)
    consolidated_output_filename = f"{output_dir}/consolidated_brand_report_{date_range}.xlsx"
    with pd.ExcelWriter(consolidated_output_filename) as writer:
        consolidated_df.to_excel(writer, sheet_name='Consolidated_Summary', index=False)

    # Apply formatting using openpyxl
    workbook = load_workbook(consolidated_output_filename)
    sheet = workbook['Consolidated_Summary']
    
    # Apply row height adjustment
    for column in sheet.columns:
        max_length = max([len(str(cell.value)) for cell in column])
        adjusted_width = max_length
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width + 1
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 17  # Adjust the value to your desired height

    for cell in sheet["1:1"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    workbook.save(consolidated_output_filename)

print("Individual brand reports and consolidated report have been successfully saved with the current date and date range in the filenames.")
