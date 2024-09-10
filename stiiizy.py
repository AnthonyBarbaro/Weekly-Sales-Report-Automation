import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Function to process each file
def process_file(file_path, output_prefix):
    if not os.path.exists(file_path):
        print(f"Error: The file at path {file_path} does not exist.")
        return

    # Load the Excel file into a DataFrame, specifying the correct header row
    df = pd.read_excel(file_path, header=4)  # Adjust this if necessary based on your file structure

    # Strip any leading/trailing spaces from column names
    df.columns = df.columns.str.strip()

    # Rename columns to handle unnamed columns properly
    df.columns = [
        "Order ID", "Order Time", "Budtender Name", "Customer Name", "Customer Type", 
        "Vendor Name", "Product Name", "Category", "Package ID", "Batch ID", 
        "External Package ID", "Total Inventory Sold", "Unit Weight Sold", "Total Weight Sold", 
        "Gross Sales", "Inventory Cost", "Discounted Amount", "Loyalty as Discount", 
        "Net Sales", "Return Date", "UPC GTIN (Canada)", "Provincial SKU (Canada)", 
        "Producer", "Order Profit"
    ]

    # Remove unnecessary columns
    columns_to_remove = [
        "Budtender Name", "Customer Type", "Package ID", "Batch ID", 
        "External Package ID", "Unit Weight Sold", 
        "Loyalty as Discount", "UPC GTIN (Canada)", "Producer"
    ]
    df = df.drop(columns=columns_to_remove)

    # Function to determine the day of the week for a given date
    def get_day_of_week(date):
        return date.strftime('%A')

    # Filter rows where Vendor Name is "Elevation (Stiiizy)"
    df_filtered = df[df['Vendor Name'] == 'Elevation (Stiiizy)'].copy()

    if df_filtered.empty:
        print(f"No data for 'Elevation (Stiiizy)' in {file_path}")
        return

    # Create a new column for the day of the week
    df_filtered['Day of Week'] = df_filtered['Order Time'].apply(get_day_of_week)

    # Filter out only the relevant days (Thursday, Friday, Saturday)
    relevant_days = ['Sunday','Monday','Tuesday','Wednesday','Thursday', 'Friday', 'Saturday']
    df_filtered = df_filtered[df_filtered['Day of Week'].isin(relevant_days)]

    # Apply category-specific filtering for each day
    thursday_categories = ["Eighths", "Pre-Rolls", "Flower", "Halves", "Quarters", "Ounces"]
    saturday_categories = ["Disposables", "Cartridges"]
    
    
    saturday_data = df_filtered[(df_filtered['Day of Week'] == 'Saturday')]
    friday_data = df_filtered[df_filtered['Day of Week'] == 'Friday']
    thursday_data = df_filtered[(df_filtered['Day of Week'] == 'Thursday')]
    wednesday_data = df_filtered[(df_filtered['Day of Week'] == 'Wednesday')]
    tuesday_data = df_filtered[(df_filtered['Day of Week'] == 'Tuesday')]
    monday_data = df_filtered[(df_filtered['Day of Week'] == 'Monday')]
    sunday_data = df_filtered[(df_filtered['Day of Week'] == 'Sunday')]

    # Combine filtered data back together for analytics
    analytics_df = pd.concat([  sunday_data,monday_data,tuesday_data,wednesday_data,thursday_data, friday_data, saturday_data])

    # Calculate the total inventory cost and 30% of the sum
    total_inventory_cost = analytics_df['Inventory Cost'].sum()
    cost_30_percent = total_inventory_cost * 0.30

    # Break down the sums for each day
    daily_sums = analytics_df.groupby('Day of Week')['Inventory Cost'].sum()

    # Most sold products with units and weight
    most_sold_products = analytics_df.groupby('Product Name').agg(
        Total_Weight_Sold=('Total Weight Sold', 'sum'),
        Total_Units_Sold=('Total Inventory Sold', 'sum')
    ).sort_values(by='Total_Weight_Sold', ascending=False).head(10)

    # Most sold categories with units and weight
    most_sold_categories = analytics_df.groupby('Category').agg(
        Total_Weight_Sold=('Total Weight Sold', 'sum'),
        Total_Units_Sold=('Total Inventory Sold', 'sum')
    ).sort_values(by='Total_Weight_Sold', ascending=False).head(10)

    # Calculate percentages for categories
    category_percentages = (analytics_df.groupby('Category')['Total Weight Sold'].sum() / analytics_df['Total Weight Sold'].sum()) * 100

    # Get the date range
    start_date = analytics_df['Order Time'].min()
    end_date = analytics_df['Order Time'].max()
    date_range = f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    
    expected_units = cost_30_percent / 14.5

    # Create a summary DataFrame
    summary_data = {
        'Total Inventory Cost': [total_inventory_cost],
        '30% of Total Cost': [cost_30_percent],
        'Expected Units (30% / 14.5)': [expected_units],
        'Date Range': [date_range]
    }
    summary_df = pd.DataFrame(summary_data)
    
    # Add daily sums to the summary
    for day, sum_cost in daily_sums.items():
        summary_df[day] = [sum_cost]

    # Most sold products for Cartridges and Disposables
    cartridges_disposables = analytics_df[analytics_df['Category'].isin(['Cartridges', 'Disposables'])]
    most_sold_cd_products = cartridges_disposables.groupby(['Category', 'Product Name']).agg(
        Total_Weight_Sold=('Total Weight Sold', 'sum'),
        Total_Units_Sold=('Total Inventory Sold', 'sum')
    ).sort_values(by='Total_Weight_Sold', ascending=False).head(10).reset_index()

    # Create detailed analysis DataFrames
    most_sold_products_df = most_sold_products.reset_index()
    most_sold_products_df.columns = ['Product Name', 'Total Weight Sold', 'Total Units Sold']

    most_sold_categories_df = most_sold_categories.reset_index()
    most_sold_categories_df.columns = ['Category', 'Total Weight Sold', 'Total Units Sold']

    category_percentages_df = category_percentages.reset_index()
    category_percentages_df.columns = ['Category', 'Percentage of Total Weight Sold']

    # Combine all DataFrames into a single report DataFrame
    report_df = pd.DataFrame()

    report_df = pd.concat([report_df, pd.DataFrame([['Summary']], columns=['Summary'])], ignore_index=True)
    report_df = pd.concat([report_df, summary_df], ignore_index=True)

    report_df = pd.concat([report_df, pd.DataFrame([['']], columns=['Summary'])], ignore_index=True)
    report_df = pd.concat([report_df, pd.DataFrame([['Most Sold Products']], columns=['Summary'])], ignore_index=True)
    report_df = pd.concat([report_df, most_sold_products_df], ignore_index=True)

    report_df = pd.concat([report_df, pd.DataFrame([['']], columns=['Summary'])], ignore_index=True)
    report_df = pd.concat([report_df, pd.DataFrame([['Most Sold Categories']], columns=['Summary'])], ignore_index=True)
    report_df = pd.concat([report_df, most_sold_categories_df], ignore_index=True)

    report_df = pd.concat([report_df, pd.DataFrame([['']], columns=['Summary'])], ignore_index=True)
    report_df = pd.concat([report_df, pd.DataFrame([['Category Percentages']], columns=['Summary'])], ignore_index=True)
    report_df = pd.concat([report_df, category_percentages_df], ignore_index=True)

    report_df = pd.concat([report_df, pd.DataFrame([['']], columns=['Summary'])], ignore_index=True)
    report_df = pd.concat([report_df, pd.DataFrame([['Most Sold Cartridges and Disposables']], columns=['Summary'])], ignore_index=True)
    report_df = pd.concat([report_df, most_sold_cd_products], ignore_index=True)

    # Generate the output file name
    output_filename = f"{output_prefix}_{date_range.replace(' ', '_').replace(':', '')}.xlsx"
    with pd.ExcelWriter(output_filename) as writer:
        analytics_df.to_excel(writer, sheet_name='Sorted Data', index=False)
        report_df.to_excel(writer, sheet_name='Analytics Report', index=False)



    # Apply formatting using openpyxl
    workbook = load_workbook(output_filename)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for column in sheet.columns:
            max_length = max([len(str(cell.value)) for cell in column])
            adjusted_width = max_length
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width + 2
        for row in sheet.iter_rows():
            sheet.row_dimensions[row[0].row].height = 17  # Adjust the value to your desired height
        for cell in sheet["1:1"]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    # Save the formatted workbook
    workbook.save(output_filename)

    print(f"Report successfully saved to {output_filename}")

# Process both files
process_file('salesLM.xlsx', 'LM')
process_file('salesMV.xlsx', 'MV')
